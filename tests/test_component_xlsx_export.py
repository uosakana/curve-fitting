import unittest
from io import BytesIO

from openpyxl import load_workbook

from app_services.export_service import component_workbook_bytes


class ComponentXlsxExportTests(unittest.TestCase):
    def test_component_workbook_contains_series_and_metadata(self):
        content = component_workbook_bytes(
            {
                "point_count": 2,
                "columns": [
                    {"key": "voltage", "label": "Voltage(V)"},
                    {"key": "measured", "label": "Measured_Current(A)"},
                ],
                "data": {
                    "voltage": [-0.5, 0.0],
                    "measured": [1e-9, 2e-9],
                },
            },
            {"sample_id": "sample-1", "structure": "ITO / ZnO / PbS"},
        )

        workbook = load_workbook(BytesIO(content), data_only=True)
        self.assertIn("components", workbook.sheetnames)
        self.assertIn("metadata", workbook.sheetnames)
        sheet = workbook["components"]
        self.assertEqual(sheet["A1"].value, "Voltage(V)")
        self.assertEqual(sheet["B1"].value, "Measured_Current(A)")
        self.assertEqual(sheet["A2"].value, -0.5)
        self.assertEqual(sheet["B3"].value, 2e-9)
        metadata = workbook["metadata"]
        self.assertEqual(metadata["A2"].value, "sample_id")
        self.assertEqual(metadata["B2"].value, "sample-1")


if __name__ == "__main__":
    unittest.main()
